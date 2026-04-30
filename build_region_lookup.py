"""
Run this script once to convert references/ebird_regions_and_codes.xls
into references/region_lookup.json — a fast lookup file used at runtime.
"""

import json
import shutil
import tempfile
from pathlib import Path

import openpyxl


EXCEL_PATH = Path("references/ebird_regions_and_codes.xls")
OUTPUT_PATH = Path("references/region_lookup.json")


def load_workbook(path: Path):
    # openpyxl requires a .xlsx extension; the file is xlsx internally
    # despite its .xls name, so we copy it to a temp file first.
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        shutil.copy(path, tmp.name)
        return openpyxl.load_workbook(tmp.name, read_only=True, data_only=True)


def sheet_to_rows(sheet) -> list[tuple]:
    rows = list(sheet.iter_rows(values_only=True))
    header, data = rows[0], rows[1:]
    return header, data


def build_countries(sheet) -> dict:
    _, rows = sheet_to_rows(sheet)
    # Each row: (country_code, country_name, continent)
    return {
        code: {"name": name, "continent": continent}
        for code, name, continent in rows
        if code  # skip any blank rows
    }


def build_subnational1(sheet) -> dict:
    _, rows = sheet_to_rows(sheet)
    # Each row: (country_code, country_name, subnational1_code, subnational1_name)
    return {
        sub1_code: sub1_name
        for _, _, sub1_code, sub1_name in rows
        if sub1_code
    }


def build_subnational2(sheet) -> dict:
    _, rows = sheet_to_rows(sheet)
    # Each row: (country_code, country_name, sub1_code, sub1_name, sub2_code, sub2_name)
    return {
        sub2_code: sub2_name
        for _, _, _, _, sub2_code, sub2_name in rows
        if sub2_code
    }


def main():
    print(f"Reading {EXCEL_PATH} ...")
    wb = load_workbook(EXCEL_PATH)

    lookup = {
        "countries":    build_countries(wb["countries"]),
        "subnational1": build_subnational1(wb["subnational1 regions"]),
        "subnational2": build_subnational2(wb["subnational2 regions"]),
    }

    OUTPUT_PATH.write_text(json.dumps(lookup, indent=2, ensure_ascii=False))

    print(f"Written to {OUTPUT_PATH}")
    print(f"  countries:    {len(lookup['countries'])} entries")
    print(f"  subnational1: {len(lookup['subnational1'])} entries")
    print(f"  subnational2: {len(lookup['subnational2'])} entries")


if __name__ == "__main__":
    main()
